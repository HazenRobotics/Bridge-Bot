import discord
import threading
import time
import openpyxl
from datetime import date
from slack_sdk import WebClient
from slack_bolt import App
from slack_bolt.adapter.socket_mode import SocketModeHandler
from discord.ext import tasks

help = ["!msg (text) messages other app",
        "!find (item) finds item in sorting system",
        "!read (word(slack only)) looks through said archive",
        "!wakeup (user)(slack only) sends user 25 DMS",
        "!mass(ask C1nner for the other half) Does a very cool thing"
        ]

token_app = 'xapp-1-A041WRBKX8E-4363930560688-aaf6a085e6d4b893b8c4221515fc1e6438d88b949de2e885609eeddc5468c8bd'
token_slack = 'xoxb-1032863990657-4071370518627-ZVecpFlXu3VTED9t3mdX9Dby'
token_discord = 'MTAxOTI0MzU0NzA0OTg3MzUyOQ.G9un4F.gg3bGB-MWefDciRUCfYEvF3KgDrQYmkBoowZyE'

slack = WebClient(token=token_slack)
app = App(token=token_slack, signing_secret="")

intents = discord.Intents.default()
intents.message_content = True

client = discord.Client(intents=intents)

target = ""
def commands(msg, bool, channel):
    global target
    if "!msg" in msg:
        message(msg[4:], not bool, channel)
        return False
    elif "!read" in msg:
        lines = ""
        with open(channel) as file:
            for line in file:
                if msg[5:] in line:
                    lines += line + "\n"
            message(line, bool, channel)
        return False
    elif "!wakeup" in msg and bool:
        target = msg[10:len(msg)-1]
        message("WAKEUP STARTED", bool, channel)
        threading.Thread(target=wakeup).start()
    elif "!help" in msg:
        helper = ""
        for h in help:
            helper += h + "\n"
        message(helper, bool, channel)
        return False
    elif "!find" in msg:
        wrkbk = openpyxl.load_workbook("sorting.xlsx")
        sh = wrkbk.active
        found = False
        for i in range(2, sh.max_row + 1):
            for j in range(1, sh.max_column + 1):
                if msg[6:] in str(sh.cell(row=i, column=j).value):
                    message(sh.cell(row=i, column=j).value + " can be found in " + sh.cell(row=i, column=1).value, bool,
                            channel)
                    found = True
                if not found:
                    message("NOT FOUND", bool, channel)
    elif "!patch" in msg:
        message("1.1 added patch multiThread wakeup Updated find", bool, channel)
    elif "<!everyone>" in msg and bool:
        message("New announcement in slack", not bool, channel)
    return True


def message(str, bool, channel):
    if bool:
        slack.chat_postMessage(channel=channel, text=str)
    else:
        channel = client.get_channel(912762550679142442)
        client.loop.create_task(channel.send(str))


@client.event
async def on_ready():
    print('discord ready')
    printer.start()


@client.event
async def on_message(message):
    if message.author == client.user:
        return
    commands(message.content, False, "#bridge-channel")


@app.event('message')
def print_message(event):
    if commands(event['text'], True, event['channel']):
        file = open(event['channel'], "a")
        file.write("[" + str(date.today()) + "]<@" + event['user'] + ">: " + event['text'] + "\n")
        print("[" + str(date.today()) + "]<@" + event['user'] + ">: " + event['text'] + "\n")
        file.close()


def wakeup():
    global target
    if bool:
        i = 25
        while i > 0:
            message("WAKE UP", True, target)
            time.sleep(0.5)
            i -= 1


@tasks.loop(seconds=1)
async def printer():
    return


def run():
    SocketModeHandler(app, token_app).start()


t = threading.Thread(target=run, args=())
t.start()
client.run(token_discord)
